from __future__ import annotations

import json
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_WORKBOOK = (
    r"\\10.150.0.179\정보사업본부\마케팅폴더(정보사업본부)\1.수시 확인 자료"
    r"\2026년 신규, 대체, 해지 유형 분석 자료(주간업무보고용).xlsx"
)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y.%m.%d")
    if isinstance(value, date):
        return value.strftime("%Y.%m.%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return " ".join(str(value).replace("\n", " ").split()).strip()


def clean_number(value: Any) -> int | float | str:
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return int(value) if float(value).is_integer() else float(value)
    text = clean_text(value).replace(",", "")
    try:
        number = float(text)
    except ValueError:
        return clean_text(value)
    return int(number) if number.is_integer() else number


def row_texts(ws, row: int, start: int = 1, end: int | None = None) -> list[str]:
    end = end or ws.max_column
    return [clean_text(ws.cell(row, column).value) for column in range(start, end + 1)]


def pair_row_values(ws, header_row: int, value_row: int, start: int = 1, end: int | None = None) -> list[dict[str, Any]]:
    end = end or ws.max_column
    items: list[dict[str, Any]] = []
    for column in range(start, end + 1):
        label = clean_text(ws.cell(header_row, column).value)
        if not label:
            continue
        value = clean_number(ws.cell(value_row, column).value)
        items.append({"label": label, "value": value})
    return items


def sparse_pair_row_values(ws, header_row: int, value_row: int, columns: list[int]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for column in columns:
        label = clean_text(ws.cell(header_row, column).value)
        if not label:
            continue
        items.append({"label": label, "value": clean_number(ws.cell(value_row, column).value)})
    return items


def source_as_of(ws) -> str:
    for row in range(1, min(ws.max_row, 8) + 1):
        for column in range(1, min(ws.max_column, 5) + 1):
            text = clean_text(ws.cell(row, column).value)
            if "기준" in text:
                return text.strip("()")
    return ""


def industry_summary(ws, start_row: int, end_row: int, columns: list[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in range(start_row, end_row + 1):
        label = clean_text(ws.cell(row, 1).value)
        if not label:
            continue
        item: dict[str, Any] = {"label": label}
        for index, key in enumerate(columns, start=2):
            item[key] = clean_number(ws.cell(row, index).value)
        rows.append(item)
    return rows


def sparse_industry_summary(ws, start_row: int, end_row: int, columns: list[tuple[int, str]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in range(start_row, end_row + 1):
        label = clean_text(ws.cell(row, 1).value)
        if not label:
            continue
        item: dict[str, Any] = {"label": label}
        for column, key in columns:
            item[key] = clean_number(ws.cell(row, column).value)
        rows.append(item)
    return rows


def first_flag(flags: dict[str, Any], fallback: str = "") -> str:
    for key, value in flags.items():
        if isinstance(value, (int, float)) and value:
            return key
        if isinstance(value, str) and value.strip():
            return key
    return fallback


def trailing_note(ws, row: int, start_col: int) -> str:
    notes = []
    for column in range(start_col, ws.max_column + 1):
        text = clean_text(ws.cell(row, column).value)
        if text:
            notes.append(text)
    return " / ".join(dict.fromkeys(notes))


def parse_new_replacement(ws) -> dict[str, Any]:
    business_labels = ["외환", "주식·선물·옵션", "채권", "기타"]
    replacement_labels = ["체크", "마켓포인트", "블룸버그", "로이터", "한경머니·기타", "신규"]
    records: list[dict[str, Any]] = []
    group = ""
    for row in range(22, ws.max_row + 1):
        first = clean_text(ws.cell(row, 1).value)
        if first.startswith("(") and first.endswith(")"):
            group = first.strip("()")
            continue
        id_code = clean_text(ws.cell(row, 3).value)
        company = clean_text(ws.cell(row, 4).value)
        if not id_code or not company or id_code in {"아이디", "ID"}:
            continue
        business_flags = {
            label: clean_number(ws.cell(row, 7 + index).value)
            for index, label in enumerate(business_labels)
        }
        replacement_flags = {
            label: clean_number(ws.cell(row, 11 + index).value)
            for index, label in enumerate(replacement_labels)
        }
        records.append(
            {
                "no": clean_number(ws.cell(row, 1).value),
                "date": clean_text(ws.cell(row, 2).value),
                "idCode": id_code,
                "companyName": company,
                "departmentName": clean_text(ws.cell(row, 5).value),
                "recommender": clean_text(ws.cell(row, 6).value),
                "businessType": first_flag(business_flags),
                "businessFlags": business_flags,
                "replacementType": first_flag(replacement_flags, "신규"),
                "replacementFlags": replacement_flags,
                "note": trailing_note(ws, row, 17),
                "group": group,
            }
        )

    return {
        "title": clean_text(ws.cell(1, 1).value),
        "asOf": source_as_of(ws),
        "workSummary": sparse_pair_row_values(ws, 4, 5, [3, 4, 5, 6, 7]),
        "replacementSummary": sparse_pair_row_values(ws, 7, 8, [3, 4, 5, 6, 7, 10, 13]),
        "industrySummary": sparse_industry_summary(
            ws,
            12,
            20,
            [
                (3, "check"),
                (4, "marketPoint"),
                (5, "bloomberg"),
                (6, "reuters"),
                (7, "hankyungEtc"),
                (10, "new"),
                (13, "total"),
            ],
        ),
        "records": records,
    }


def parse_termination(ws) -> dict[str, Any]:
    reason_labels = [
        "사용자퇴사·이직",
        "비용절감·예산삭감",
        "활용도저조·불필요",
        "콘텐츠불만·타사대체",
        "조직개편·업무변경",
        "휴직·장기출장",
        "회사합병매각",
        "계약만료",
        "구독료 미수",
    ]
    competitor_labels = ["체크", "마켓포인트", "블룸버그", "로이터", "한경·기타", "아웃"]
    records: list[dict[str, Any]] = []
    group = ""
    for row in range(22, ws.max_row + 1):
        first = clean_text(ws.cell(row, 1).value)
        if first.startswith("(") and first.endswith(")"):
            group = first.strip("()")
            continue
        id_code = clean_text(ws.cell(row, 3).value)
        company = clean_text(ws.cell(row, 4).value)
        if not id_code or not company or id_code in {"아이디", "ID"}:
            continue
        reason_flags = {
            label: clean_number(ws.cell(row, 7 + index).value)
            for index, label in enumerate(reason_labels)
        }
        competitor_flags = {
            label: clean_number(ws.cell(row, 16 + index).value)
            for index, label in enumerate(competitor_labels)
        }
        records.append(
            {
                "no": clean_number(ws.cell(row, 1).value),
                "date": clean_text(ws.cell(row, 2).value),
                "idCode": id_code,
                "companyName": company,
                "departmentName": clean_text(ws.cell(row, 5).value),
                "recommender": clean_text(ws.cell(row, 6).value),
                "reason": first_flag(reason_flags),
                "reasonFlags": reason_flags,
                "competitorType": first_flag(competitor_flags, "아웃"),
                "competitorFlags": competitor_flags,
                "note": trailing_note(ws, row, 22),
                "penalty": clean_number(ws.cell(row, 27).value),
                "group": group,
            }
        )

    return {
        "title": clean_text(ws.cell(1, 1).value),
        "asOf": source_as_of(ws),
        "reasonSummary": sparse_pair_row_values(ws, 4, 5, [3, 4, 5, 6, 7, 11, 14, 19, 23, 24]),
        "competitorSummary": sparse_pair_row_values(ws, 7, 8, [3, 4, 5, 6, 7, 14, 19]),
        "industrySummary": sparse_industry_summary(
            ws,
            12,
            20,
            [
                (3, "userMove"),
                (4, "costCut"),
                (5, "lowUsage"),
                (6, "contentOrCompetitor"),
                (7, "contractEnd"),
                (11, "reorg"),
                (14, "leave"),
                (19, "merger"),
                (23, "unpaid"),
                (24, "total"),
            ],
        ),
        "records": records,
    }


AREA_NET_GROWTH_LABELS = [
    "국내은행/지주",
    "국내증권",
    "외국계은행, 외국계증권",
    "자산운용",
    "보험사",
    "일반기업,대학교",
    "공제회, 중개사, 선물사, 공사, 개인 등 기타금융 전체",
    "연기금, 공기업, 정부",
]


def normalize_area_net_growth_group(value: Any) -> str:
    text = clean_text(value).replace("해지", "").strip()
    compact = text.replace(" ", "")
    if not compact:
        return AREA_NET_GROWTH_LABELS[6]
    if "국내은행" in compact or "지주" in compact:
        return AREA_NET_GROWTH_LABELS[0]
    if "국내증권" in compact:
        return AREA_NET_GROWTH_LABELS[1]
    if "외국계" in compact or "외국" in compact:
        return AREA_NET_GROWTH_LABELS[2]
    if "자산운용" in compact or "운용" in compact:
        return AREA_NET_GROWTH_LABELS[3]
    if "보험" in compact:
        return AREA_NET_GROWTH_LABELS[4]
    if "연기금" in compact or "정부" in compact or "공기업" in compact or "협회" in compact or "금감" in compact:
        return AREA_NET_GROWTH_LABELS[7]
    if "대학교" in compact or "대학" in compact or "학교" in compact or "일반기업" in compact or "기업체" in compact:
        return AREA_NET_GROWTH_LABELS[5]
    return AREA_NET_GROWTH_LABELS[6]


def parse_area_net_growth(ws) -> dict[str, Any]:
    replacement_labels = ["체크", "마켓포인트", "블룸버그", "로이터", "한경머니·기타", "신규"]
    termination_labels = [
        "사용자퇴사·이직",
        "비용절감·예산삭감",
        "활용도저조·불필요",
        "콘텐츠불만·타사대체",
        "조직개편·업무변경",
        "휴직·장기출장",
        "회사합병매각",
        "계약만료",
        "구독료 미수",
    ]
    summary_rows: list[dict[str, Any]] = []
    for row in range(4, 13):
        label = clean_text(ws.cell(row, 1).value)
        if not label:
            continue
        summary_rows.append(
            {
                "no": clean_text(ws.cell(row, 1).value),
                "area": clean_text(ws.cell(row, 2).value),
                "manager": clean_text(ws.cell(row, 5).value),
                "newCount": clean_number(ws.cell(row, 7).value),
                "terminationCount": clean_number(ws.cell(row, 9).value),
                "netCount": clean_number(ws.cell(row, 11).value),
            }
        )

    records: list[dict[str, Any]] = []
    group = ""
    record_kind = "new"
    for row in range(14, ws.max_row + 1):
        first = clean_text(ws.cell(row, 1).value)
        if first.startswith("(") and first.endswith(")"):
            group_text = first.strip("()")
            record_kind = "termination" if "해지" in group_text else "new"
            group = normalize_area_net_growth_group(group_text)
            continue
        id_code = clean_text(ws.cell(row, 3).value)
        company = clean_text(ws.cell(row, 4).value)
        if (
            not id_code
            or not company
            or id_code in {"아이디", "ID", "0"}
            or company == "0"
            or id_code == "00:00:00"
            or company == "00:00:00"
            or first in {"구분", "소 계", "총 계"}
        ):
            continue
        base_record = {
            "no": clean_number(ws.cell(row, 1).value),
            "date": clean_text(ws.cell(row, 2).value),
            "idCode": id_code,
            "companyName": company,
            "departmentName": clean_text(ws.cell(row, 5).value),
            "note": clean_text(ws.cell(row, 6).value) if record_kind == "termination" else trailing_note(ws, row, 13),
            "group": group,
            "areaGroup": group,
        }
        if record_kind == "termination":
            reason_flags = {
                label: clean_number(ws.cell(row, 7 + index).value)
                for index, label in enumerate(termination_labels)
            }
            if not any(isinstance(value, (int, float)) and value for value in reason_flags.values()):
                continue
            records.append(
                {
                    **base_record,
                    "kind": "termination",
                    "transactionType": "해지",
                    "recommender": "",
                    "reason": first_flag(reason_flags, "계약만료"),
                    "reasonFlags": reason_flags,
                }
            )
        else:
            replacement_flags = {
                label: clean_number(ws.cell(row, 7 + index).value)
                for index, label in enumerate(replacement_labels)
            }
            if not any(isinstance(value, (int, float)) and value for value in replacement_flags.values()):
                continue
            records.append(
                {
                    **base_record,
                    "kind": "new",
                    "transactionType": "신규/대체",
                    "recommender": clean_text(ws.cell(row, 6).value),
                    "replacementType": first_flag(replacement_flags, "신규"),
                    "replacementFlags": replacement_flags,
                }
            )

    return {
        "title": clean_text(ws.cell(1, 1).value),
        "asOf": source_as_of(ws) or clean_text(ws.cell(1, 1).value).split("(")[-1].strip(")") if "(" in clean_text(ws.cell(1, 1).value) else "",
        "summaryRows": summary_rows,
        "records": records,
    }


def parse_personal(ws) -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    for row in range(4, ws.max_row + 1):
        manager = clean_text(ws.cell(row, 2).value)
        if not manager:
            continue
        rows.append(
            {
                "no": clean_text(ws.cell(row, 1).value),
                "manager": manager,
                "totalNew": clean_number(ws.cell(row, 3).value),
                "new": clean_number(ws.cell(row, 4).value),
                "check": clean_number(ws.cell(row, 5).value),
                "marketPoint": clean_number(ws.cell(row, 6).value),
                "reutersBloomberg": clean_number(ws.cell(row, 7).value),
            }
        )
    return {
        "title": clean_text(ws.cell(1, 1).value),
        "asOf": source_as_of(ws) or clean_text(ws.cell(1, 1).value).split("(")[-1].strip(")") if "(" in clean_text(ws.cell(1, 1).value) else "",
        "rows": rows,
    }


def main() -> int:
    workbook_path = Path(sys.argv[1] if len(sys.argv) > 1 else DEFAULT_WORKBOOK)
    output_path = Path(sys.argv[2] if len(sys.argv) > 2 else "data/type-analysis-source.json")
    wb = load_workbook(workbook_path, data_only=True)
    payload = {
        "version": 1,
        "sourceFile": str(workbook_path),
        "sourceUpdatedAt": datetime.fromtimestamp(workbook_path.stat().st_mtime).isoformat(),
        "extractedAt": datetime.now().isoformat(),
        "excludedSheets": ["최근 5개년도 업종별 순증추이", "Sheet1"],
        "newReplacement": parse_new_replacement(wb["2026년 신규,대체 분석"]),
        "terminationType": parse_termination(wb["2026년 해지 유형 분석"]),
        "areaNetGrowth": parse_area_net_growth(wb["2026년 영역별 순증 분석"]),
        "personalPerformance": parse_personal(wb["2026년 개인별 실적"]),
        "weeklySnapshots": [],
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(
        json.dumps(
            {
                "newRecords": len(payload["newReplacement"]["records"]),
                "terminationRecords": len(payload["terminationType"]["records"]),
                "areaRecords": len(payload["areaNetGrowth"]["records"]),
                "personalRows": len(payload["personalPerformance"]["rows"]),
                "output": str(output_path),
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
